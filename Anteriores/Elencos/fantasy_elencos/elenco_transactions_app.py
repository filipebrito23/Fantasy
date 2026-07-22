from pathlib import Path
import json
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from data_loader import load_workbook_data, SEASONS
from transforms import (
    SEASON_LABELS,
    get_team_options,
    get_visible_seasons,
    build_roster_view,
    format_roster_for_display,
    calculate_main_totals,
    calculate_dev_totals,
)

st.set_page_config(page_title="Elencos Fantasy NBA", layout="wide")
st.title("Elencos Fantasy NBA")
st.caption("Versão com cadastro de transactions no app e gravação no workbook.")

DEFAULT_FILE = Path("roster.xlsx")
TX_SHEET = "transactions"
TX_ITEMS_SHEET = "transaction_times"

@st.cache_data
def cached_load(file_path: str):
    return load_workbook_data(file_path)


def currency(v: float) -> str:
    if pd.isna(v):
        return "-"
    return f"US$ {v:,.2f}"


def display_table(df: pd.DataFrame):
    st.dataframe(df, use_container_width=True, hide_index=True)


def ensure_sheet_headers(ws, headers):
    existing = [c.value for c in ws[1]]
    if not existing or all(v is None for v in existing):
        ws.append(headers)
        return
    if list(existing) != list(headers):
        raise ValueError(f"Cabeçalhos divergentes em {ws.title}")


def get_next_id(df: pd.DataFrame, col: str) -> int:
    if df.empty or col not in df.columns or df[col].dropna().empty:
        return 1
    return int(pd.to_numeric(df[col], errors="coerce").max()) + 1


def append_transaction_to_workbook(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)
    if TX_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TX_SHEET)
        ensure_sheet_headers(ws, list(tx_row.keys()))
    else:
        ws = wb[TX_SHEET]
        ensure_sheet_headers(ws, list(tx_row.keys()))
    ws.append([tx_row.get(h) for h in tx_row.keys()])

    if TX_ITEMS_SHEET not in wb.sheetnames:
        ws_items = wb.create_sheet(TX_ITEMS_SHEET)
        ensure_sheet_headers(ws_items, list(item_rows[0].keys()) if item_rows else [])
    else:
        ws_items = wb[TX_ITEMS_SHEET]
        if item_rows:
            ensure_sheet_headers(ws_items, list(item_rows[0].keys()))
    for row in item_rows:
        ws_items.append([row.get(h) for h in item_rows[0].keys()])
    wb.save(file_path)


def build_red_flags(df: pd.DataFrame, visible_seasons: list[str]) -> pd.DataFrame:
    out = df.copy()
    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        opt_col = f"TO_{season}"
        if sal_col not in out.columns or opt_col not in out.columns:
            continue
        flag_col = f"{sal_col}__red"
        out[flag_col] = out[opt_col].astype(str).str.strip().str.lower().eq("sim")
        out = out.drop(columns=[opt_col])
    return out


def render_red_table(df: pd.DataFrame, visible_seasons: list[str]) -> str:
    out = df.copy()
    drop_cols = [c for c in out.columns if str(c).startswith("TO_")]
    if drop_cols:
        out = out.drop(columns=drop_cols)
    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        flag_col = f"{sal_col}__red"
        if sal_col in out.columns and flag_col in out.columns:
            out[sal_col] = out[sal_col].apply(lambda x: f"{x}" if pd.notna(x) else "")
            out = out.drop(columns=[flag_col])
    return out.to_html(escape=False, index=False)


def is_player_asset(item_type: str) -> bool:
    return str(item_type).strip().lower() in {"player", "jogador"}


def apply_transactions(roster_df: pd.DataFrame, tx_items: pd.DataFrame, team_id: int, roster_type: str) -> pd.DataFrame:
    if roster_df.empty or tx_items.empty:
        return roster_df
    out = roster_df.copy()
    rel = tx_items.copy()
    if "from_roster_type" in rel.columns:
        rel = rel[rel["from_roster_type"].astype(str).str.upper().eq(roster_type)]
    if "from_team_id" in rel.columns:
        rel = rel[rel["from_team_id"] == team_id]
    if rel.empty:
        return out
    player_moves = rel[rel["item_type"].apply(is_player_asset)] if "item_type" in rel.columns else rel.iloc[0:0]
    if not player_moves.empty and "player_id" in out.columns and "asset_id" in player_moves.columns:
        removed = set(player_moves.loc[player_moves["to_team_id"].ne(team_id), "asset_id"].dropna().tolist())
        if removed:
            out = out[~out["player_id"].isin(removed)]
        added = player_moves.loc[player_moves["to_team_id"].eq(team_id), "asset_id"].dropna().tolist()
        if added:
            existing = set(out["player_id"].tolist())
            extras = roster_df[roster_df["player_id"].isin(added) & ~roster_df["player_id"].isin(existing)]
            if not extras.empty:
                out = pd.concat([out, extras], ignore_index=True)
    return out


def to_ws_date(value):
    if pd.isna(value) or value is None:
        return None
    return pd.to_datetime(value).to_pydatetime()


if not DEFAULT_FILE.exists():
    st.error("Arquivo roster.xlsx não encontrado na pasta do projeto.")
    st.stop()

if "saved_tx" not in st.session_state:
    st.session_state.saved_tx = False

data = cached_load(str(DEFAULT_FILE))
teams = get_team_options(data["teams"])
team_map = data["teams"][["team_id", "team_name"]].drop_duplicates() if not data["teams"].empty else pd.DataFrame(columns=["team_id", "team_name"])
team_lookup = dict(zip(team_map["team_id"], team_map["team_name"])) if not team_map.empty else {}
player_lookup = dict(zip(data["players"]["player_id"], data["players"]["player_name"])) if not data["players"].empty else {}
player_options = data["players"][["player_id", "player_name"]].dropna().copy()
player_options["label"] = player_options["player_id"].astype(str) + " - " + player_options["player_name"].astype(str)

c1, c2 = st.columns([2, 1])
with c1:
    selected_team_name = st.selectbox("Selecione o time", teams["team_name"].tolist())
with c2:
    selected_start_season = st.selectbox("Temporada inicial", SEASONS, format_func=lambda x: SEASON_LABELS[x])

selected_team_id = int(teams.loc[teams["team_name"] == selected_team_name, "team_id"].iloc[0])
visible_seasons = get_visible_seasons(selected_start_season)

st.subheader("Nova transaction")
with st.form("tx_form", clear_on_submit=True):
    tx_type = st.selectbox("Tipo", ["TRADE", "WAIVER", "SIGN", "CUT", "OTHER"])
    tx_date = st.date_input("Data da transaction")
    from_team_name = st.selectbox("Time origem", teams["team_name"].tolist(), index=teams["team_name"].tolist().index(selected_team_name))
    to_team_name = st.selectbox("Time destino", teams["team_name"].tolist())
    initiated_by = st.selectbox("Iniciada por", ["FROM", "TO", "BOTH", "LEAGUE"])
    status = st.selectbox("Status", ["DRAFT", "APPROVED", "PENDING"])
    effective_season = st.selectbox("Temporada efetiva", SEASONS, format_func=lambda x: SEASON_LABELS[x])
    notes = st.text_area("Observações")
    st.markdown("Itens")
    n_items = st.number_input("Quantidade de itens", min_value=1, max_value=6, value=1, step=1)
    item_rows = []
    for i in range(int(n_items)):
        cols = st.columns([2, 2, 2, 2])
        with cols[0]:
            item_type = st.selectbox(f"Item tipo {i+1}", ["player", "pick"], key=f"item_type_{i}")
        with cols[1]:
            asset_id = st.selectbox(f"Ativo {i+1}", player_options["label"].tolist(), key=f"asset_{i}") if item_type == "player" else st.text_input(f"Ativo {i+1}", key=f"asset_{i}")
        with cols[2]:
            from_roster_type = st.selectbox(f"Roster origem {i+1}", ["MAIN", "DEV"], key=f"from_rt_{i}")
        with cols[3]:
            to_roster_type = st.selectbox(f"Roster destino {i+1}", ["MAIN", "DEV"], key=f"to_rt_{i}")
        item_rows.append((item_type, asset_id, from_roster_type, to_roster_type))
    submitted = st.form_submit_button("Salvar transaction")

if submitted:
    try:
        txs = data.get("transactions", pd.DataFrame())
        tx_items_df = data.get("transaction_times", pd.DataFrame())
        tx_id = get_next_id(txs, "transaction_id")
        tx_row = {
            "transaction_id": tx_id,
            "transaction_type": tx_type,
            "transaction_date": to_ws_date(tx_date),
            "season": effective_season,
            "from_team_id": int(teams.loc[teams["team_name"] == from_team_name, "team_id"].iloc[0]),
            "to_team_id": int(teams.loc[teams["team_name"] == to_team_name, "team_id"].iloc[0]),
            "initiated_by": initiated_by,
            "status": status,
            "notes": notes,
        }
        item_payload = []
        next_item_id = get_next_id(tx_items_df, "item_id")
        for idx, (item_type, asset_id, from_rt, to_rt) in enumerate(item_rows, start=0):
            if item_type == "player":
                pid = int(str(asset_id).split(" - ")[0])
            else:
                pid = asset_id
            item_payload.append({
                "transaction_id": tx_id,
                "item_id": next_item_id + idx,
                "item_type": item_type,
                "asset_id": pid,
                "from_team_id": int(teams.loc[teams["team_name"] == from_team_name, "team_id"].iloc[0]),
                "to_team_id": int(teams.loc[teams["team_name"] == to_team_name, "team_id"].iloc[0]),
                "from_roster_type": from_rt,
                "to_roster_type": to_rt,
                "effective_season": effective_season,
                "item_notes": notes,
            })
        append_transaction_to_workbook(str(DEFAULT_FILE), tx_row, item_payload)
        st.success("Transaction salva no workbook com sucesso.")
        st.cache_data.clear()
        st.session_state.saved_tx = True
        st.rerun()
    except Exception as e:
        st.error(f"Erro ao salvar transaction: {e}")

transactions = data.get("transactions", pd.DataFrame())
transaction_times = data.get("transaction_times", pd.DataFrame())
transactions_view = transactions.copy() if not transactions.empty else pd.DataFrame()
transaction_times_view = transaction_times.copy() if not transaction_times.empty else pd.DataFrame()
if not transactions_view.empty:
    if "from_team_id" in transactions_view.columns:
        transactions_view["from_team"] = transactions_view["from_team_id"].map(team_lookup)
    if "to_team_id" in transactions_view.columns:
        transactions_view["to_team"] = transactions_view["to_team_id"].map(team_lookup)
if not transaction_times_view.empty:
    if "from_team_id" in transaction_times_view.columns:
        transaction_times_view["from_team"] = transaction_times_view["from_team_id"].map(team_lookup)
    if "to_team_id" in transaction_times_view.columns:
        transaction_times_view["to_team"] = transaction_times_view["to_team_id"].map(team_lookup)
    if "asset_id" in transaction_times_view.columns:
        transaction_times_view["asset_name"] = transaction_times_view["asset_id"].map(player_lookup).fillna(transaction_times_view["asset_id"].astype(str))

main_team_df = data["roster"].loc[data["roster"]["team_id"] == selected_team_id].copy()
dev_team_df = data["development"].loc[data["development"]["team_id"] == selected_team_id].copy()
main_team_tx = apply_transactions(main_team_df, transaction_times, selected_team_id, "MAIN")
dev_team_tx = apply_transactions(dev_team_df, transaction_times, selected_team_id, "DEV")

main_roster_raw = build_roster_view(main_team_tx, data["players"], selected_team_id, "MAIN", visible_seasons)
main_roster = format_roster_for_display(main_roster_raw, visible_seasons)
main_totals = calculate_main_totals(main_team_tx, data["fines"], selected_team_id, visible_seasons)

dev_roster_raw = build_roster_view(dev_team_tx, data["players"], selected_team_id, "DEV", visible_seasons)
dev_roster = format_roster_for_display(dev_roster_raw, visible_seasons)
dev_totals = calculate_dev_totals(dev_team_tx, visible_seasons)

with st.expander("Diagnóstico de carregamento", expanded=False):
    st.json({
        "players": len(data["players"]),
        "teams": len(data["teams"]),
        "roster": len(data["roster"]),
        "development": len(data["development"]),
        "picks": len(data["picks"]),
        "fines": len(data["fines"]),
        "transactions": len(transactions),
        "transaction_times": len(transaction_times),
        "time_selecionado": selected_team_name,
        "temporadas_visiveis": [SEASON_LABELS[s] for s in visible_seasons],
    })

st.subheader("Transações")
if transactions_view.empty:
    st.info("Não há transações cadastradas ainda.")
else:
    cols_tx = [c for c in ["transaction_id", "transaction_type", "transaction_date", "season", "from_team", "to_team", "initiated_by", "status", "notes"] if c in transactions_view.columns]
    display_table(transactions_view[cols_tx])

st.subheader("Itens por transação")
if transaction_times_view.empty:
    st.info("Nenhum item vinculado às transações ainda.")
else:
    cols_ti = [c for c in ["transaction_id", "item_id", "item_type", "asset_name", "from_team", "to_team", "from_roster_type", "to_roster_type", "effective_season", "item_notes"] if c in transaction_times_view.columns]
    display_table(transaction_times_view[cols_ti])

st.subheader("Elenco principal")
display_main = build_red_flags(main_roster, visible_seasons)
for season in visible_seasons:
    label = SEASON_LABELS[season]
    if label in display_main.columns:
        display_main[label] = display_main[label].apply(currency)
option_cols_main = [c for c in display_main.columns if str(c).startswith("TO ")]
for col in option_cols_main:
    salary_col = col.replace("TO ", "")
    if salary_col in display_main.columns:
        mask = display_main[col].astype(str).str.strip().str.lower().eq("sim")
        display_main.loc[mask, salary_col] = display_main.loc[mask, salary_col].map(lambda x: f"🔴 {x}")
        display_main.loc[mask, col] = display_main.loc[mask, col].map(lambda x: f"🔴 {x}")
display_table(display_main)

st.subheader("Totalizadores do elenco principal")
main_totals_display = main_totals.copy()
for col in ["Salários", "Multas", "Cap restante"]:
    if col in main_totals_display.columns:
        main_totals_display[col] = main_totals_display[col].apply(currency)
st.dataframe(main_totals_display, use_container_width=True, hide_index=True)

st.subheader("Liga de desenvolvimento")
display_dev = build_red_flags(dev_roster, visible_seasons)
for season in visible_seasons:
    label = SEASON_LABELS[season]
    if label in display_dev.columns:
        display_dev[label] = display_dev[label].apply(currency)
option_cols_dev = [c for c in display_dev.columns if str(c).startswith("TO ")]
for col in option_cols_dev:
    salary_col = col.replace("TO ", "")
    if salary_col in display_dev.columns:
        mask = display_dev[col].astype(str).str.strip().str.lower().eq("sim")
        display_dev.loc[mask, salary_col] = display_dev.loc[mask, salary_col].map(lambda x: f"🔴 {x}")
        display_dev.loc[mask, col] = display_dev.loc[mask, col].map(lambda x: f"🔴 {x}")
display_table(display_dev)

st.subheader("Totalizadores da development")
dev_totals_display = dev_totals.copy()
for col in ["Salários", "Cap restante"]:
    if col in dev_totals_display.columns:
        dev_totals_display[col] = dev_totals_display[col].apply(currency)
st.dataframe(dev_totals_display, use_container_width=True, hide_index=True)
