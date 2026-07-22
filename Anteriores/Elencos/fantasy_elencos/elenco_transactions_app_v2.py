from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from data_loader import load_workbook_data, SEASONS
from transforms import (
    SEASON_LABELS,
    get_team_options,
    get_visible_seasons,
    build_roster_view,
    format_roster_for_display,
    calculate_main_totals,
    calculate_dev_totals,
)

st.set_page_config(page_title="Elencos Fantasy NBA", layout="wide")
st.title("Elencos Fantasy NBA")
st.caption("Transactions com domínio por time e aplicação imediata no workbook.")

DEFAULT_FILE = Path("roster.xlsx")
TX_SHEET = "transactions"
TX_ITEMS_SHEET = "transaction_times"

@st.cache_data
def cached_load(file_path: str):
    return load_workbook_data(file_path)


def currency(v: float) -> str:
    if pd.isna(v):
        return "-"
    return f"US$ {v:,.2f}"


def display_table(df: pd.DataFrame):
    st.dataframe(df, use_container_width=True, hide_index=True)


def ensure_sheet_headers(ws, headers):
    existing = [c.value for c in ws[1]]
    if not existing or all(v is None for v in existing):
        ws.append(headers)
        return
    if list(existing) != list(headers):
        raise ValueError(f"Cabeçalhos divergentes em {ws.title}")


def get_next_id(df: pd.DataFrame, col: str) -> int:
    if df.empty or col not in df.columns or df[col].dropna().empty:
        return 1
    return int(pd.to_numeric(df[col], errors="coerce").max()) + 1


def append_rows(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)
    if TX_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TX_SHEET)
        ws.append(list(tx_row.keys()))
    else:
        ws = wb[TX_SHEET]
        ensure_sheet_headers(ws, list(tx_row.keys()))
    ws.append([tx_row.get(h) for h in tx_row.keys()])

    if item_rows:
        if TX_ITEMS_SHEET not in wb.sheetnames:
            ws_items = wb.create_sheet(TX_ITEMS_SHEET)
            ws_items.append(list(item_rows[0].keys()))
        else:
            ws_items = wb[TX_ITEMS_SHEET]
            ensure_sheet_headers(ws_items, list(item_rows[0].keys()))
        for row in item_rows:
            ws_items.append([row.get(h) for h in item_rows[0].keys()])
    wb.save(file_path)


def load_sheet_df(wb, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    data = list(ws.values)
    if not data:
        return pd.DataFrame()
    return pd.DataFrame(data[1:], columns=data[0])


def save_sheet_df(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    if df.empty:
        return
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([row.get(c) for c in df.columns])


def is_player_asset(item_type: str) -> bool:
    return str(item_type).strip().lower() in {"player", "jogador"}


def team_domain_ids(data, team_id: int, roster_type: str) -> tuple[set, set]:
    if roster_type == "MAIN":
        roster = data["roster"]
    else:
        roster = data["development"]
    roster_ids = set(roster.loc[roster["team_id"] == team_id, "player_id"].dropna().astype(int).tolist()) if not roster.empty else set()
    picks = data.get("picks", pd.DataFrame())
    pick_ids = set()
    if not picks.empty:
        owner_cols = [c for c in picks.columns if "owner" in c.lower() or "team" in c.lower()]
        if owner_cols:
            owner_col = owner_cols[0]
            pick_ids = set(picks.loc[picks[owner_col] == team_id, picks.columns[0]].astype(str).tolist())
    return roster_ids, pick_ids


def validate_items(data, from_team_id: int, item_rows: list[dict]) -> list[str]:
    errors = []
    main_ids, main_picks = team_domain_ids(data, from_team_id, "MAIN")
    dev_ids, _ = team_domain_ids(data, from_team_id, "DEV")
    for i, item in enumerate(item_rows, start=1):
        item_type = item["item_type"]
        asset_id = item["asset_id"]
        from_rt = item["from_roster_type"]
        if item_type == "player":
            pid = int(asset_id)
            allowed = main_ids if from_rt == "MAIN" else dev_ids
            if pid not in allowed:
                errors.append(f"Item {i}: jogador não pertence ao domínio do time na roster {from_rt}.")
        elif item_type == "pick":
            if str(asset_id) not in main_picks:
                errors.append(f"Item {i}: pick não pertence ao domínio do time.")
    return errors


def apply_transaction_to_rosters(data, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(DEFAULT_FILE)
    roster_df = load_sheet_df(wb, "roster")
    dev_df = load_sheet_df(wb, "development")
    picks_df = load_sheet_df(wb, "picks")

    from_team_id = int(tx_row["from_team_id"])
    to_team_id = int(tx_row["to_team_id"])

    for item in item_rows:
        if item["item_type"] == "player":
            pid = int(item["asset_id"])
            if item["from_roster_type"] == "MAIN":
                mask_from = (roster_df["team_id"].astype(int) == from_team_id) & (roster_df["player_id"].astype(int) == pid)
                mask_to = (roster_df["team_id"].astype(int) == to_team_id) & (roster_df["player_id"].astype(int) == pid)
                if mask_from.any():
                    roster_df.loc[mask_from, "team_id"] = to_team_id
                elif not mask_to.any():
                    row = roster_df[(roster_df["player_id"].astype(int) == pid)].head(1).copy()
                    if not row.empty:
                        row.loc[:, "team_id"] = to_team_id
                        roster_df = pd.concat([roster_df, row], ignore_index=True)
            else:
                mask_from = (dev_df["team_id"].astype(int) == from_team_id) & (dev_df["player_id"].astype(int) == pid)
                if mask_from.any():
                    dev_df.loc[mask_from, "team_id"] = to_team_id
        elif item["item_type"] == "pick":
            if not picks_df.empty:
                pid = str(item["asset_id"])
                pick_cols = [c for c in picks_df.columns if "pick" in c.lower() or "id" == c.lower()]
                if pick_cols:
                    id_col = pick_cols[0]
                    owner_cols = [c for c in picks_df.columns if "owner" in c.lower() or "team" in c.lower()]
                    if owner_cols:
                        owner_col = owner_cols[0]
                        mask = picks_df[id_col].astype(str).eq(pid) & picks_df[owner_col].astype(int).eq(from_team_id)
                        picks_df.loc[mask, owner_col] = to_team_id

    save_sheet_df(wb, "roster", roster_df)
    save_sheet_df(wb, "development", dev_df)
    save_sheet_df(wb, "picks", picks_df)
    if TX_SHEET in wb.sheetnames:
        txs = load_sheet_df(wb, TX_SHEET)
        if not txs.empty:
            save_sheet_df(wb, TX_SHEET, txs)
    if TX_ITEMS_SHEET in wb.sheetnames:
        tx_items = load_sheet_df(wb, TX_ITEMS_SHEET)
        if not tx_items.empty:
            save_sheet_df(wb, TX_ITEMS_SHEET, tx_items)
    wb.save(DEFAULT_FILE)


def build_red_flags(df: pd.DataFrame, visible_seasons: list[str]) -> pd.DataFrame:
    out = df.copy()
    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        opt_col = f"TO_{season}"
        if sal_col not in out.columns or opt_col not in out.columns:
            continue
        out[f"{sal_col}__red"] = out[opt_col].astype(str).str.strip().str.lower().eq("sim")
        out = out.drop(columns=[opt_col])
    return out


def render_red_table(df: pd.DataFrame, visible_seasons: list[str]) -> str:
    out = df.copy()
    drop_cols = [c for c in out.columns if str(c).startswith("TO_")]
    if drop_cols:
        out = out.drop(columns=drop_cols)
    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        flag_col = f"{sal_col}__red"
        if sal_col in out.columns and flag_col in out.columns:
            out[sal_col] = out[sal_col].apply(lambda x: f"{x}" if pd.notna(x) else "")
            out = out.drop(columns=[flag_col])
    return out.to_html(escape=False, index=False)

if not DEFAULT_FILE.exists():
    st.error("Arquivo roster.xlsx não encontrado na pasta do projeto.")
    st.stop()

if "saved_tx" not in st.session_state:
    st.session_state.saved_tx = False

data = cached_load(str(DEFAULT_FILE))
teams = get_team_options(data["teams"])
team_map = data["teams"][["team_id", "team_name"]].drop_duplicates() if not data["teams"].empty else pd.DataFrame(columns=["team_id", "team_name"])
team_lookup = dict(zip(team_map["team_id"], team_map["team_name"])) if not team_map.empty else {}
player_lookup = dict(zip(data["players"]["player_id"], data["players"]["player_name"])) if not data["players"].empty else {}
player_options = data["players"][["player_id", "player_name"]].dropna().copy()
player_options["label"] = player_options["player_id"].astype(str) + " - " + player_options["player_name"].astype(str)
pick_options = data.get("picks", pd.DataFrame()).copy()
if not pick_options.empty:
    pick_options["label"] = pick_options.iloc[:, 0].astype(str)

c1, c2 = st.columns([2, 1])
with c1:
    selected_team_name = st.selectbox("Selecione o time", teams["team_name"].tolist())
with c2:
    selected_start_season = st.selectbox("Temporada inicial", SEASONS, format_func=lambda x: SEASON_LABELS[x])

selected_team_id = int(teams.loc[teams["team_name"] == selected_team_name, "team_id"].iloc[0])
visible_seasons = get_visible_seasons(selected_start_season)

st.subheader("Nova transaction")
with st.form("tx_form", clear_on_submit=True):
    tx_type = st.selectbox("Tipo", ["TRADE", "WAIVER", "SIGN", "CUT", "DLEAGUE", "OTHER"])
    tx_date = st.date_input("Data da transaction")
    from_team_name = st.selectbox("Time origem", teams["team_name"].tolist(), index=teams["team_name"].tolist().index(selected_team_name))
    to_team_name = st.selectbox("Time destino", teams["team_name"].tolist())
    initiated_by = st.selectbox("Iniciada por", ["FROM", "TO", "BOTH", "LEAGUE"])
    status = st.selectbox("Status", ["DRAFT", "APPROVED", "PENDING"])
    effective_season = st.selectbox("Temporada efetiva", SEASONS, format_func=lambda x: SEASON_LABELS[x])
    notes = st.text_area("Observações")
    st.markdown("Itens")
    n_items = st.number_input("Quantidade de itens", min_value=1, max_value=6, value=1, step=1)
    item_rows = []
    from_team_id = int(teams.loc[teams["team_name"] == from_team_name, "team_id"].iloc[0])
    main_ids, pick_ids = team_domain_ids(data, from_team_id, "MAIN")
    dev_ids, _ = team_domain_ids(data, from_team_id, "DEV")
    main_player_choices = player_options[player_options["player_id"].isin(main_ids)]["label"].tolist() if not player_options.empty else []
    dev_player_choices = player_options[player_options["player_id"].isin(dev_ids)]["label"].tolist() if not player_options.empty else []
    pick_choices = pick_options[pick_options.iloc[:, 0].astype(str).isin(pick_ids)]["label"].tolist() if not pick_options.empty else []
    for i in range(int(n_items)):
        cols = st.columns([2, 2, 2, 2])
        with cols[0]:
            item_type = st.selectbox(f"Item tipo {i+1}", ["player", "pick"], key=f"item_type_{i}")
        with cols[1]:
            if item_type == "player":
                roster_type = st.selectbox(f"Roster origem {i+1}", ["MAIN", "DEV"], key=f"from_rt_{i}")
                choices = main_player_choices if roster_type == "MAIN" else dev_player_choices
                asset_choice = st.selectbox(f"Ativo {i+1}", choices if choices else ["Sem ativos disponíveis"], key=f"asset_{i}")
            else:
                roster_type = st.selectbox(f"Roster origem {i+1}", ["MAIN"], key=f"from_rt_{i}")
                asset_choice = st.selectbox(f"Ativo {i+1}", pick_choices if pick_choices else ["Sem picks disponíveis"], key=f"asset_{i}")
        with cols[2]:
            to_roster_type = st.selectbox(f"Roster destino {i+1}", ["MAIN", "DEV"], key=f"to_rt_{i}")
        with cols[3]:
            item_note = st.text_input(f"Nota {i+1}", key=f"item_note_{i}")
        if item_type == "player":
            pid = int(str(asset_choice).split(" - ")[0]) if " - " in str(asset_choice) else -1
        else:
            pid = str(asset_choice).split(" - ")[0]
        item_rows.append({"item_type": item_type, "asset_id": pid, "from_roster_type": roster_type, "to_roster_type": to_roster_type, "item_notes": item_note})
    submitted = st.form_submit_button("Salvar transaction")

if submitted:
    try:
        txs = data.get("transactions", pd.DataFrame())
        tx_items_df = data.get("transaction_times", pd.DataFrame())
        tx_id = get_next_id(txs, "transaction_id")
        tx_row = {
            "transaction_id": tx_id,
            "transaction_type": tx_type,
            "transaction_date": pd.to_datetime(tx_date).to_pydatetime(),
            "season": effective_season,
            "from_team_id": from_team_id,
            "to_team_id": int(teams.loc[teams["team_name"] == to_team_name, "team_id"].iloc[0]),
            "initiated_by": initiated_by,
            "status": status,
            "notes": notes,
        }
        errs = validate_items(data, from_team_id, item_rows)
        if errs:
            st.error(" | ".join(errs))
            st.stop()
        next_item_id = get_next_id(tx_items_df, "item_id")
        payload = []
        for idx, item in enumerate(item_rows, start=0):
            row = {
                "transaction_id": tx_id,
                "item_id": next_item_id + idx,
                "item_type": item["item_type"],
                "asset_id": item["asset_id"],
                "from_team_id": from_team_id,
                "to_team_id": int(teams.loc[teams["team_name"] == to_team_name, "team_id"].iloc[0]),
                "from_roster_type": item["from_roster_type"],
                "to_roster_type": item["to_roster_type"],
                "effective_season": effective_season,
                "item_notes": item["item_notes"],
            }
            payload.append(row)
        append_rows(str(DEFAULT_FILE), tx_row, payload)
        apply_transaction_to_rosters(data, tx_row, payload)
        st.success("Transaction salva e aplicada no workbook com sucesso.")
        st.cache_data.clear()
        st.rerun()
    except Exception as e:
        st.error(f"Erro ao salvar transaction: {e}")

transactions = data.get("transactions", pd.DataFrame())
transaction_times = data.get("transaction_times", pd.DataFrame())
transactions_view = transactions.copy() if not transactions.empty else pd.DataFrame()
transaction_times_view = transaction_times.copy() if not transaction_times.empty else pd.DataFrame()
if not transactions_view.empty:
    if "from_team_id" in transactions_view.columns:
        transactions_view["from_team"] = transactions_view["from_team_id"].map(team_lookup)
    if "to_team_id" in transactions_view.columns:
        transactions_view["to_team"] = transactions_view["to_team_id"].map(team_lookup)
if not transaction_times_view.empty:
    if "from_team_id" in transaction_times_view.columns:
        transaction_times_view["from_team"] = transaction_times_view["from_team_id"].map(team_lookup)
    if "to_team_id" in transaction_times_view.columns:
        transaction_times_view["to_team"] = transaction_times_view["to_team_id"].map(team_lookup)
    if "asset_id" in transaction_times_view.columns:
        transaction_times_view["asset_name"] = transaction_times_view["asset_id"].map(player_lookup).fillna(transaction_times_view["asset_id"].astype(str))

main_team_df = data["roster"].loc[data["roster"]["team_id"] == selected_team_id].copy()
dev_team_df = data["development"].loc[data["development"]["team_id"] == selected_team_id].copy()
main_team_tx = main_team_df.copy()
dev_team_tx = dev_team_df.copy()

main_roster_raw = build_roster_view(main_team_tx, data["players"], selected_team_id, "MAIN", visible_seasons)
main_roster = format_roster_for_display(main_roster_raw, visible_seasons)
main_totals = calculate_main_totals(main_team_tx, data["fines"], selected_team_id, visible_seasons)

dev_roster_raw = build_roster_view(dev_team_tx, data["players"], selected_team_id, "DEV", visible_seasons)
dev_roster = format_roster_for_display(dev_roster_raw, visible_seasons)
dev_totals = calculate_dev_totals(dev_team_tx, visible_seasons)

with st.expander("Diagnóstico de carregamento", expanded=False):
    st.json({
        "players": len(data["players"]),
        "teams": len(data["teams"]),
        "roster": len(data["roster"]),
        "development": len(data["development"]),
        "picks": len(data["picks"]),
        "fines": len(data["fines"]),
        "transactions": len(transactions),
        "transaction_times": len(transaction_times),
        "time_selecionado": selected_team_name,
        "temporadas_visiveis": [SEASON_LABELS[s] for s in visible_seasons],
    })

st.subheader("Transações")
if transactions_view.empty:
    st.info("Não há transações cadastradas ainda.")
else:
    cols_tx = [c for c in ["transaction_id", "transaction_type", "transaction_date", "season", "from_team", "to_team", "initiated_by", "status", "notes"] if c in transactions_view.columns]
    display_table(transactions_view[cols_tx])

st.subheader("Itens por transação")
if transaction_times_view.empty:
    st.info("Nenhum item vinculado às transações ainda.")
else:
    cols_ti = [c for c in ["transaction_id", "item_id", "item_type", "asset_name", "from_team", "to_team", "from_roster_type", "to_roster_type", "effective_season", "item_notes"] if c in transaction_times_view.columns]
    display_table(transaction_times_view[cols_ti])

st.subheader("Elenco principal")
display_main = build_red_flags(main_roster, visible_seasons)
for season in visible_seasons:
    label = SEASON_LABELS[season]
    if label in display_main.columns:
        display_main[label] = display_main[label].apply(currency)
option_cols_main = [c for c in display_main.columns if str(c).startswith("TO ")]
for col in option_cols_main:
    salary_col = col.replace("TO ", "")
    if salary_col in display_main.columns:
        mask = display_main[col].astype(str).str.strip().str.lower().eq("sim")
        display_main.loc[mask, salary_col] = display_main.loc[mask, salary_col].map(lambda x: f"🔴 {x}")
        display_main.loc[mask, col] = display_main.loc[mask, col].map(lambda x: f"🔴 {x}")
display_table(display_main)

st.subheader("Totalizadores do elenco principal")
main_totals_display = main_totals.copy()
for col in ["Salários", "Multas", "Cap restante"]:
    if col in main_totals_display.columns:
        main_totals_display[col] = main_totals_display[col].apply(currency)
st.dataframe(main_totals_display, use_container_width=True, hide_index=True)

st.subheader("Liga de desenvolvimento")
display_dev = build_red_flags(dev_roster, visible_seasons)
for season in visible_seasons:
    label = SEASON_LABELS[season]
    if label in display_dev.columns:
        display_dev[label] = display_dev[label].apply(currency)
option_cols_dev = [c for c in display_dev.columns if str(c).startswith("TO ")]
for col in option_cols_dev:
    salary_col = col.replace("TO ", "")
    if salary_col in display_dev.columns:
        mask = display_dev[col].astype(str).str.strip().str.lower().eq("sim")
        display_dev.loc[mask, salary_col] = display_dev.loc[mask, salary_col].map(lambda x: f"🔴 {x}")
        display_dev.loc[mask, col] = display_dev.loc[mask, col].map(lambda x: f"🔴 {x}")
display_table(display_dev)

st.subheader("Totalizadores da development")
dev_totals_display = dev_totals.copy()
for col in ["Salários", "Cap restante"]:
    if col in dev_totals_display.columns:
        dev_totals_display[col] = dev_totals_display[col].apply(currency)
st.dataframe(dev_totals_display, use_container_width=True, hide_index=True)
